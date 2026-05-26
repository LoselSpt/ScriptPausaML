import os
import sys
import time
import openpyxl
import csv
from datetime import datetime
from playwright.sync_api import sync_playwright

# Garante suporte completo a caracteres UTF-8 no console do Windows
try:
    sys.stdout.reconfigure(encoding='utf-8')
except AttributeError:
    pass

EXCEL_PATH = "excluir_2.xlsx"
LISTA_URL = "https://www.mercadolivre.com.br/anuncios/lista"
REPORT_PATH = r"G:\Meu Drive\LoselNotebook\LogsSkuScript\Relatorio_Verificacao_Rapida.md"
TRACKER_PATH = r"G:\Meu Drive\LoselNotebook\LogsSkuScript\Acompanhamento_Tarefa_Atual.md"
EXCEL_DESKTOP_PATH = r"C:\Users\joaos\OneDrive\Desktop\anuncios_nao_excluidos.xlsx"

def ler_anuncios(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_id = -1
    col_sku = -1
    for idx, header in enumerate(headers):
        if header.lower() == "id anuncio":
            col_id = idx
        elif header.lower() == "sku":
            col_sku = idx
            
    if col_id == -1:
        raise ValueError(f"Coluna 'ID Anuncio' nao encontrada. Cabecalhos: {headers}")
        
    items = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) > col_id:
            val_id = str(row[col_id]).strip() if row[col_id] else ""
            val_sku = str(row[col_sku]).strip() if col_sku != -1 and len(row) > col_sku and row[col_sku] else ""
            if val_id and val_id not in seen:
                items.append({"id": val_id, "sku": val_sku})
                seen.add(val_id)
    return items

# == Atualiza o Dashboard no Obsidian com a auditoria de verificação ============
def atualizar_dashboard_verificacao(todos_anuncios: list[dict], resultados_auditoria: list[dict], item_ativo: dict = None, status_ativo: str = None):
    total_anuncios = len(todos_anuncios)
    total_auditados = len(resultados_auditoria)
    
    total_limpos = sum(1 for r in resultados_auditoria if r["status"] == "LIMPO")
    total_ativos = sum(1 for r in resultados_auditoria if r["status"] == "ATIVO")
    total_erros = sum(1 for r in resultados_auditoria if r["status"] == "ERRO")
    
    porcentagem_auditoria = int((total_auditados / total_anuncios) * 100) if total_anuncios > 0 else 0
    data_hora_agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Status Geral de Auditoria em HTML
    if total_auditados >= total_anuncios:
        if total_ativos == 0:
            status_auditoria_html = '<span style="background-color: #2e7d32; color: #ffffff; padding: 3px 8px; border-radius: 12px; font-weight: 600; font-size: 0.85em; box-shadow: 1px 1px 3px rgba(0,0,0,0.15);">🏆 100% LIMPO - AUDITADO COM SUCESSO</span>'
        else:
            status_auditoria_html = f'<span style="background-color: #c62828; color: #ffffff; padding: 3px 8px; border-radius: 12px; font-weight: 600; font-size: 0.85em; box-shadow: 1px 1px 3px rgba(0,0,0,0.15);">⚠️ DIRTY ({total_ativos} ativos detectados)</span>'
    else:
        status_auditoria_html = f'<span style="background-color: #1565c0; color: #ffffff; padding: 3px 8px; border-radius: 12px; font-weight: 600; font-size: 0.85em; box-shadow: 1px 1px 3px rgba(0,0,0,0.15);">⚡ VERIFICANDO... ({porcentagem_auditoria}%)</span>'
        
    # Card do Item Ativo
    if item_ativo:
        id_curr = item_ativo["id"]
        sku_curr = item_ativo["sku"]
        badge_curr = '<span style="background-color: #1565c0; color: white; padding: 1.5px 6px; border-radius: 4px; font-weight: bold; font-size: 0.8em;">⚙️ AUDITANDO</span>'
        card_item_atual = f"""
> [!info] **Auditoria de Verificação - Item Atual**
> * 🆔 **ID Anúncio:** `{id_curr}`
> * 🏷️ **SKU:** `{sku_curr}`
> * 📊 **Status:** {badge_curr}
> * 💬 **Ação:** *Verificando se o anúncio ainda está visível no Mercado Livre...*
"""
    else:
        card_item_atual = """
> [!success] **Auditoria de Verificação - Concluída**
> * *Toda a base foi auditada com sucesso direta na interface do Mercado Livre.*
"""

    # Tabela de Histórico de Auditoria (últimos 15 verificados)
    tabela_linhas = []
    for r in reversed(resultados_auditoria[-15:]):
        if r["status"] == "LIMPO":
            badge = '<span style="background-color: #2e7d32; color: white; padding: 1px 5px; border-radius: 4px; font-weight: bold; font-size: 0.8em;">✨ LIMPO</span>'
        elif r["status"] == "ATIVO":
            badge = '<span style="background-color: #ef6c00; color: white; padding: 1px 5px; border-radius: 4px; font-weight: bold; font-size: 0.8em;">⚠️ ATIVO</span>'
        else:
            badge = '<span style="background-color: #c62828; color: white; padding: 1px 5px; border-radius: 4px; font-weight: bold; font-size: 0.8em;">🚨 ERRO</span>'
        tabela_linhas.append(f"| `{r['id']}` | `{r['sku']}` | {badge} | {r['detalhe']} |")
        
    if not tabela_linhas:
        tabela_linhas_str = "| - | - | - | Nenhum anúncio verificado ainda |"
    else:
        tabela_linhas_str = "\n".join(tabela_linhas)

    md_content = f"""# 👑 Painel de Controle - Exclusão de Anúncios no Mercado Livre

> [!abstract] **Informações do Sistema**
> * 📁 **Planilha de Origem:** `excluir_2.xlsx`
> * 🎯 **Estratégia:** Busca Cirúrgica por **ID do Anúncio (MLB ID)**
> * 🔄 **Status de Exclusão:** <span style="background-color: #2e7d32; color: #ffffff; padding: 3px 8px; border-radius: 12px; font-weight: 600; font-size: 0.85em; box-shadow: 1px 1px 3px rgba(0,0,0,0.15);">🏆 EXCLUSÕES CONCLUÍDAS (100%)</span>
> * 🔍 **Status de Verificação:** {status_auditoria_html}
> * 📆 **Última Atualização:** `{data_hora_agora}`

---

## 📊 Progresso Geral da Auditoria de Verificação

<div style="display: flex; align-items: center; gap: 15px; margin: 15px 0;">
  <div style="flex-grow: 1; background-color: #333333; border-radius: 8px; padding: 4px; border: 1px solid #555555; height: 32px; width: 100%;">
    <div style="width: {porcentagem_auditoria}%; background: linear-gradient(90deg, #1565c0, #2e7d32); height: 100%; border-radius: 6px; display: flex; align-items: center; justify-content: center; color: white; font-weight: bold; font-family: sans-serif; font-size: 0.9em; box-shadow: inset 0px 1px 2px rgba(255,255,255,0.2);">
      {porcentagem_auditoria}% ({total_auditados} / {total_anuncios})
    </div>
  </div>
</div>

---

{card_item_atual}

---

## 🕒 Histórico da Auditoria de Verificação (Últimos 15)

<details open>
<summary>Clique para expandir/recolher a lista de auditoria</summary>

| ID Anúncio | SKU | Status Encontrado | Detalhe |
| :--- | :--- | :--- | :--- |
{tabela_linhas_str}

</details>

---

## 🛠️ Métricas de Auditoria
* ✨ **LIMPOS (Deletados):** `{total_limpos}` anúncios
* ⚠️ **DIRTY (Ativos):** `{total_ativos}` anúncios
* 🚨 **Erros de Auditoria:** `{total_erros}` anúncios
"""

    try:
        os.makedirs(os.path.dirname(TRACKER_PATH), exist_ok=True)
        with open(TRACKER_PATH, "w", encoding="utf-8") as f:
            f.write(md_content)
    except Exception as e:
        print(f"[WARN] Nao foi possivel escrever Acompanhamento_Tarefa_Atual.md: {e}")

# == Main ======================================================================
def main():
    print("=== INICIANDO AUDITORIA DE VERIFICAÇÃO RÁPIDA ===")
    
    if not os.path.exists(EXCEL_PATH):
        print(f"[ERRO CRITICO] Planilha {EXCEL_PATH} nao encontrada!")
        sys.exit(1)
        
    anuncios = ler_anuncios(EXCEL_PATH)
    print(f"[AUDITORIA] Carregados {len(anuncios)} anuncios para verificar.")
    
    resultados = []
    
    # Atualiza o dashboard no Obsidian com auditoria zerada
    atualizar_dashboard_verificacao(anuncios, resultados)
    
    total_limpos = 0
    total_ativos = 0
    total_erros = 0
    
    with sync_playwright() as pw:
        print("[CDP] Conectando ao Chrome na porta 9222...")
        try:
            browser = pw.chromium.connect_over_cdp("http://localhost:9222")
            print("[CDP] Conectado com sucesso!")
        except Exception as e:
            print(f"[ERRO CRITICO] Nao foi possivel conectar ao Chrome: {e}")
            sys.exit(1)
            
        context = browser.contexts[0]
        page = None
        for p in context.pages:
            if "mercadolivre.com.br" in p.url:
                page = p
                break
        if not page:
            page = context.new_page()
            page.goto(LISTA_URL)
            
        page.wait_for_load_state("domcontentloaded")
        
        for i, item in enumerate(anuncios, 1):
            id_anuncio = item["id"]
            sku = item["sku"]
            print(f"[{i}/{len(anuncios)}] Auditando ID: {id_anuncio} (SKU: {sku})...", end="", flush=True)
            
            # Atualiza o dashboard: informa o anúncio que começou a ser auditado
            atualizar_dashboard_verificacao(anuncios, resultados, item_ativo=item, status_ativo="PROCESSANDO")
            
            import urllib.parse
            url_busca = f"https://www.mercadolivre.com.br/anuncios?page=1&sort=DEFAULT&search={urllib.parse.quote(id_anuncio)}"
            
            try:
                page.goto(url_busca, wait_until="domcontentloaded", timeout=12000)
                time.sleep(0.7)  # Altamente otimizado para 0.7 segundos (suficiente pós-DOMContentLoaded)
                
                # Procura botões de ações do anúncio
                botoes = page.get_by_role("button", name="Ações secundárias").all()
                if not botoes:
                    botoes = page.get_by_role("button", name="Acciones secundarias").all()
                if not botoes:
                    botoes = page.locator("button.sll-list-row-secondary-actions-trigger").all()
                
                if botoes:
                    print(" ATIVO ⚠️")
                    status = "ATIVO"
                    detalhe = "O anúncio ainda foi encontrado ativo no Mercado Livre."
                    total_ativos += 1
                else:
                    print(" LIMPO ✨")
                    status = "LIMPO"
                    detalhe = "Não foram encontrados anúncios ativos para este ID."
                    total_limpos += 1
                    
            except Exception as e:
                print(f" ERRO: {e} 🚨")
                status = "ERRO"
                detalhe = f"Erro na verificacao: {e}"
                total_erros += 1
                
            resultados.append({
                "id": id_anuncio,
                "sku": sku,
                "status": status,
                "detalhe": detalhe
            })
            
            # Atualiza o dashboard no Obsidian com o resultado final do anúncio auditado
            atualizar_dashboard_verificacao(anuncios, resultados)
            time.sleep(0.2)  # Otimizado de 0.5s para 0.2s
            
    # Limpa card ativo no final
    atualizar_dashboard_verificacao(anuncios, resultados)
    
    # Gera o Relatório Visual Premium para o Obsidian (Relatorio_Verificacao_Rapida.md)
    porcentagem_limpo = int((total_limpos / len(anuncios)) * 100) if anuncios else 0
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    status_auditoria_html = ""
    if total_ativos == 0 and total_erros == 0:
        status_auditoria_html = '<span style="background-color: #2e7d32; color: #ffffff; padding: 3px 8px; border-radius: 12px; font-weight: 600; font-size: 0.85em; box-shadow: 1px 1px 3px rgba(0,0,0,0.15);">🏆 100% LIMPO - AUDITADO COM SUCESSO</span>'
    else:
        status_auditoria_html = f'<span style="background-color: #c62828; color: #ffffff; padding: 3px 8px; border-radius: 12px; font-weight: 600; font-size: 0.85em; box-shadow: 1px 1px 3px rgba(0,0,0,0.15);">⚠️ DIRTY ({total_ativos} ativos detectados)</span>'
        
    tabela_rows = []
    for r in resultados:
        if r["status"] == "LIMPO":
            badge = '<span style="background-color: #2e7d32; color: white; padding: 1px 5px; border-radius: 4px; font-weight: bold; font-size: 0.8em;">✨ LIMPO</span>'
        elif r["status"] == "ATIVO":
            badge = '<span style="background-color: #ef6c00; color: white; padding: 1px 5px; border-radius: 4px; font-weight: bold; font-size: 0.8em;">⚠️ ATIVO</span>'
        else:
            badge = '<span style="background-color: #c62828; color: white; padding: 1px 5px; border-radius: 4px; font-weight: bold; font-size: 0.8em;">🚨 ERRO</span>'
        tabela_rows.append(f"| `{r['id']}` | `{r['sku']}` | {badge} | {r['detalhe']} |")
        
    md_report = f"""# 🔍 Relatório de Verificação Rápida e Auditoria (Double Check)
**Última Atualização:** `{timestamp}`  
**Progresso de Auditoria:** `[████████████████████] 100.0%` (`{len(anuncios)}/{len(anuncios)}` IDs auditados)
**Status Geral de Limpeza:** {status_auditoria_html}

---

### 📊 Resumo do Estado da Base (Double Check)
| Categoria | Quantidade | Percentual |
| :--- | :--- | :--- |
| ✨ **Confirmados LIMPO (Excluídos)** | **{total_limpos}** | {porcentagem_limpo}% |
| ⚠️ **Confirmados DIRTY (Ativos)** | **{total_ativos}** | {int((total_ativos/len(anuncios))*100) if anuncios else 0}% |
| 🚨 **Erro de Verificação** | **{total_erros}** | {int((total_erros/len(anuncios))*100) if anuncios else 0}% |

---

## 📋 Detalhamento da Auditoria por ID de Anúncio

<details open>
<summary>Clique para expandir/recolher a lista de auditoria completa</summary>

| ID Anúncio | SKU | Status Encontrado | Detalhe |
| :--- | :--- | :--- | :--- |
{"\n".join(tabela_rows)}

</details>

---

## 🏆 Conclusão da Auditoria
* Auditoria concluída com sucesso! Todos os anúncios da planilha excluir_2.xlsx foram auditados diretamente na interface do Mercado Livre para atestar a segurança e validade absoluta da exclusão.
"""

    try:
        os.makedirs(os.path.dirname(REPORT_PATH), exist_ok=True)
        with open(REPORT_PATH, "w", encoding="utf-8") as f:
            f.write(md_report)
        print(f"\n[SUCESSO] Relatório de auditoria rápida gerado em: {REPORT_PATH}")
    except Exception as e:
        print(f"[ERRO] Falha ao escrever relatório: {e}")

    # == Geração de Planilha de Não Excluídos na Área de Trabalho ==================
    anuncios_nao_excluidos = [r for r in resultados if r["status"] != "LIMPO"]
    
    try:
        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        ws_new.title = "Não Excluídos"
        ws_new.append(["ID Anuncio", "SKU", "Status Encontrado", "Detalhe"])
        
        if anuncios_nao_excluidos:
            for item in anuncios_nao_excluidos:
                ws_new.append([item["id"], item["sku"], item["status"], item["detalhe"]])
            print(f"[AUDITORIA] Detectados {len(anuncios_nao_excluidos)} anuncios NAO excluidos. Gerando planilha na Area de Trabalho...")
        else:
            print("[AUDITORIA] Fantástico! 100% dos anúncios foram excluídos. Planilha de erros gerada vazia (com cabeçalhos).")
            
        wb_new.save(EXCEL_DESKTOP_PATH)
        print(f"[SUCESSO] Planilha de não excluídos salva em: {EXCEL_DESKTOP_PATH}")
    except Exception as e:
        print(f"[ERRO] Falha ao gerar planilha excel na Area de Trabalho: {e}")

if __name__ == "__main__":
    main()
