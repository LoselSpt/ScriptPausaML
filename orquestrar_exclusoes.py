import os
import sys
import subprocess
import re
import time
from datetime import datetime

# Configurações de diretórios e caminhos
SCRIPT_DIR = r"C:\Users\joaos\OneDrive\Documentos\ScriptPausa"
SCRIPT_PATH = os.path.join(SCRIPT_DIR, "excluir_anuncios_ml.py")
DRIVE_DIR = r"G:\Meu Drive\LoselNotebook\LogsSkuScript"
MAX_RODADAS = 8

# Garante suporte completo a caracteres UTF-8 no console do Windows
try:
    sys.stdout.reconfigure(encoding='utf-8')
except AttributeError:
    pass

def log_local(mensagem: str):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {mensagem}", flush=True)

def carregar_progresso_completo(drive_dir, script_dir):
    excel_path = os.path.join(script_dir, "excluir_2.xlsx")
    log_path = os.path.join(drive_dir, "status_exclusoes.csv")
    
    todos_ids = set()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_id = -1
        for idx, h in enumerate(headers):
            if h.lower() == "id anuncio":
                col_id = idx
                break
        if col_id != -1:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) > col_id and row[col_id]:
                    todos_ids.add(str(row[col_id]).strip())
    except Exception as e:
        print(f"[WARN] Nao foi possivel carregar planilha excluir_2: {e}")
        
    total_anuncios = len(todos_ids) if todos_ids else 238
    
    excluidos_concluidos = set()
    itens_falhados = set()
    if os.path.exists(log_path):
        try:
            import csv
            with open(log_path, "r", newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader, None)
                for row in reader:
                    if len(row) >= 4:
                        id_anuncio = row[1].strip()
                        status = row[3].strip()
                        if status == "CONCLUIDO":
                            excluidos_concluidos.add(id_anuncio)
                            if id_anuncio in itens_falhados:
                                itens_falhados.remove(id_anuncio)
                        elif status == "ERRO":
                            if id_anuncio not in excluidos_concluidos:
                                itens_falhados.add(id_anuncio)
        except Exception as e:
            print(f"[WARN] Nao foi possivel ler status_exclusoes.csv: {e}")
            
    return total_anuncios, len(excluidos_concluidos), len(itens_falhados)

def atualizar_acompanhamento_tempo_real(numero_rodada, total_concluidos_rodada, total_erros_rodada, drive_dir, script_dir, itens_concluidos, itens_com_erro):
    total_anuncios, total_excluidos, total_falhados = carregar_progresso_completo(drive_dir, script_dir)
    
    porcentagem = int((total_excluidos / total_anuncios) * 100) if total_anuncios > 0 else 0
    
    data_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    status_html = ""
    if total_excluidos >= total_anuncios:
        status_html = '<span style="background-color: #2e7d32; color: #ffffff; padding: 3px 8px; border-radius: 12px; font-weight: 600; font-size: 0.85em; box-shadow: 1px 1px 3px rgba(0,0,0,0.15);">🏆 CONCLUÍDO COM SUCESSO</span>'
    else:
        status_html = f'<span style="background-color: #1565c0; color: #ffffff; padding: 3px 8px; border-radius: 12px; font-weight: 600; font-size: 0.85em; box-shadow: 1px 1px 3px rgba(0,0,0,0.15);">⚡ EM EXECUÇÃO (RODADA {numero_rodada})</span>'
        
    tabela_itens = []
    for id_anuncio, info in itens_concluidos.items():
        badge = '<span style="background-color: #2e7d32; color: white; padding: 1px 5px; border-radius: 4px; font-weight: bold; font-size: 0.8em;">CONCLUÍDO</span>'
        tabela_itens.append(f"| `{id_anuncio}` | `{info['sku']}` | {badge} | {info['detalhe']} |")
        
    for id_anuncio, info in itens_com_erro.items():
        badge = '<span style="background-color: #c62828; color: white; padding: 1px 5px; border-radius: 4px; font-weight: bold; font-size: 0.8em;">FALHA / LAG</span>'
        tabela_itens.append(f"| `{id_anuncio}` | `{info['sku']}` | {badge} | {info['detalhe']} |")
        
    if not tabela_itens:
        tabela_itens_str = "| - | - | - | Nenhum anúncio processado nesta rodada |"
    else:
        tabela_itens_str = "\n".join(tabela_itens)
        
    md_content = f"""# 👑 Painel de Controle - Exclusão de Anúncios no Mercado Livre

> [!abstract] **Informações do Sistema**
> * 📁 **Planilha de Origem:** `excluir_2.xlsx`
> * 🎯 **Estratégia:** Busca Cirúrgica por **ID do Anúncio (MLB ID)**
> * 🔄 **Status Geral:** {status_html}
> * 📆 **Última Atualização:** `{data_hora}`

---

## 📊 Progresso Geral da Limpeza

<div style="display: flex; align-items: center; gap: 15px; margin: 15px 0;">
  <div style="flex-grow: 1; background-color: #333333; border-radius: 8px; padding: 4px; border: 1px solid #555555; height: 32px; width: 100%;">
    <div style="width: {porcentagem}%; background: linear-gradient(90deg, #1565c0, #2e7d32); height: 100%; border-radius: 6px; display: flex; align-items: center; justify-content: center; color: white; font-weight: bold; font-family: sans-serif; font-size: 0.9em; box-shadow: inset 0px 1px 2px rgba(255,255,255,0.2);">
      {porcentagem}% ({total_excluidos} / {total_anuncios})
    </div>
  </div>
</div>

---

## 🚀 Status da Rodada Atual ({numero_rodada})

> [!summary] **Métricas de Performance da Rodada {numero_rodada}**
> * ✅ **Concluídos com Sucesso:** `{total_concluidos_rodada}` anúncios
> * ⏳ **Falhas ou Lags a re-tentar:** `{total_erros_rodada}` anúncios
> * 🕒 **Intervalo Humano Ativo:** `3.5s carregamento` | `2.0s menu` | `3.0s pausa` (Garantindo estabilidade)

---

## 📋 Lista de Anúncios Processados na Rodada

<details open>
<summary>Clique para expandir/recolher a lista de itens da Rodada {numero_rodada}</summary>

| ID Anúncio | SKU | Status | Detalhes / Ação |
| :--- | :--- | :--- | :--- |
{tabela_itens_str}

</details>

---

## 🛠️ Próximos Passos Planejados
1. 🔄 O orquestrador executará novas rodadas subsequentes de repetição até atingir **100% de exclusão** da lista.
2. Cada rodada re-tentará somente os anúncios que falharam por lag ou delay da interface.
3. Ao finalizar tudo, um relatório consolidado com o status de cada item será gerado e salvo aqui.
"""

    caminho_tracker = os.path.join(drive_dir, "Acompanhamento_Tarefa_Atual.md")
    try:
        with open(caminho_tracker, "w", encoding="utf-8") as f:
            f.write(md_content)
        print(f"[TRACKER] Acompanhamento_Tarefa_Atual.md atualizado com sucesso no Google Drive.")
    except Exception as e:
        print(f"[WARN] Nao foi possivel escrever Acompanhamento_Tarefa_Atual.md: {e}")

def executar_rodada(numero_rodada: int) -> tuple[bool, int, int, int, str]:
    """
    Executa o script de exclusão, captura o output em tempo real e analisa os resultados.
    Retorna (sucesso, total_concluidos, total_erros, total_itens, relatorio_md)
    """
    log_local(f"Iniciando Rodada {numero_rodada}...")
    
    # Inicia o processo python de exclusão
    process = subprocess.Popen(
        [sys.executable, SCRIPT_PATH],
        cwd=SCRIPT_DIR,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding='utf-8',
        errors='replace'
    )

    stdout_lines = []
    
    # Lê a saída em tempo real e imprime no console local
    while True:
        line = process.stdout.readline()
        if not line and process.poll() is not None:
            break
        if line:
            sys.stdout.write(line)
            sys.stdout.flush()
            stdout_lines.append(line.strip())

    process.wait()
    exit_code = process.returncode
    log_local(f"Rodada {numero_rodada} finalizada com código de saída: {exit_code}")

    full_stdout = "\n".join(stdout_lines)
    
    # Variáveis de métricas da rodada
    total_itens = 0
    total_concluidos = 0
    total_erros = 0

    itens_concluidos = {}
    itens_com_erro = {}

    # Padrões Regex para análise do log
    regex_planilha = re.compile(r"\[PLANILHA\] (\d+) anuncios")
    regex_resumo_concluido = re.compile(r"Concluídos nesta rodada: (\d+)")
    regex_resumo_erros = re.compile(r"Erros/Timeouts nesta rodada: (\d+)")
    
    regex_anuncio = re.compile(r"Anúncio \d+/\d+: (.+?) \(SKU: (.+?)\)")
    regex_resumo_item = re.compile(r"\[RESUMO ITEM\] Status: (.+?) \| Detalhe: (.+)")

    current_id = None
    current_sku = None

    # Analisa linha por linha
    for line in stdout_lines:
        match_planilha = regex_planilha.search(line)
        if match_planilha:
            total_itens = int(match_planilha.group(1))
            continue
            
        match_concluidos = regex_resumo_concluido.search(line)
        if match_concluidos:
            total_concluidos = int(match_concluidos.group(1))
            continue
            
        match_erros = regex_resumo_erros.search(line)
        if match_erros:
            total_erros = int(match_erros.group(1))
            continue

        match_anuncio = regex_anuncio.search(line)
        if match_anuncio:
            current_id = match_anuncio.group(1).strip()
            current_sku = match_anuncio.group(2).strip()
            continue
            
        match_resumo_item = regex_resumo_item.search(line)
        if match_resumo_item and current_id:
            status = match_resumo_item.group(1).strip()
            detalhe = match_resumo_item.group(2).strip()
            if status == "CONCLUIDO":
                itens_concluidos[current_id] = {"sku": current_sku, "detalhe": detalhe}
            else:
                itens_com_erro[current_id] = {"sku": current_sku, "detalhe": detalhe}
            current_id = None
            current_sku = None
            continue

    # Atualiza o arquivo de progresso no Google Drive com layout Obsidian Premium
    atualizar_acompanhamento_tempo_real(
        numero_rodada, 
        total_concluidos, 
        total_erros, 
        DRIVE_DIR, 
        SCRIPT_DIR, 
        itens_concluidos, 
        itens_com_erro
    )

    # Define o status da rodada
    status_rodada = "SUCESSO COMPLETO" if (total_concluidos == 0 and total_erros == 0) else "RODANDO RE-TENTATIVAS"
    if exit_code != 0:
        status_rodada = "FALHA (CÓDIGO DE ERRO)"

    # Gera o conteúdo formatado em Markdown
    timestamp_fim = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    md_lines = [
        f"# Relatório de Execução - Rodada {numero_rodada}",
        f"**Data/Hora de Conclusão:** {timestamp_fim}  ",
        f"**Status da Rodada:** `{status_rodada}`",
        "",
        "## Resumo das Métricas",
        "| Métrica | Valor |",
        "| :--- | :--- |",
        f"| **Anúncios Totais na Planilha** | {total_itens} |",
        f"| **Excluídos / Concluídos nesta Rodada** | {total_concluidos} |",
        f"| **Falhas / Timeouts nesta Rodada** | {total_erros} |",
        f"| **Código de Saída do Script** | {exit_code} |",
        ""
    ]

    # Adiciona detalhes de anúncios excluídos
    if itens_concluidos:
        md_lines.append("## Detalhamento de Exclusões por ID de Anúncio")
        md_lines.append("| ID Anúncio | SKU | Detalhe / Resultado |")
        md_lines.append("| :--- | :--- | :--- |")
        for id_anuncio, info in itens_concluidos.items():
            md_lines.append(f"| `{id_anuncio}` | `{info['sku']}` | {info['detalhe']} |")
        md_lines.append("")

    # Adiciona detalhes de anúncios com problemas (Erros ou Timeouts)
    if itens_com_erro:
        md_lines.append("## Alertas e Problemas Identificados (Lags/Timeouts)")
        md_lines.append("| ID Anúncio | SKU | Tipo de Ocorrência |")
        md_lines.append("| :--- | :--- | :--- |")
        for id_anuncio, info in itens_com_erro.items():
            md_lines.append(f"| `{id_anuncio}` | `{info['sku']}` | `{info['detalhe']}` |")
        md_lines.append("")

    # Adiciona o Log bruto da rodada em aba retrátil
    md_lines.append("## Log Técnico da Rodada")
    md_lines.append("<details>")
    md_lines.append("<summary>Clique aqui para visualizar o Log de Saída completo</summary>")
    md_lines.append("")
    md_lines.append("```text")
    md_lines.append(full_stdout)
    md_lines.append("```")
    md_lines.append("</details>")
    md_lines.append("")

    relatorio_md = "\n".join(md_lines)
    return (exit_code == 0, total_concluidos, total_erros, total_itens, relatorio_md)

def main():
    log_local("=== INICIANDO INTEGRAÇÃO DO ORQUESTRADOR ===")
    
    if not os.path.exists(DRIVE_DIR):
        log_local(f"Criando a pasta no Google Drive: '{DRIVE_DIR}'...")
        os.makedirs(DRIVE_DIR, exist_ok=True)

    rodada = 1
    total_acumulado_excluidos = 0

    while rodada <= MAX_RODADAS:
        log_local(f"\n--- RODADA {rodada} DE {MAX_RODADAS} ---")
        
        # Executa a rodada
        sucesso, concluidos, erros, itens, relatorio = executar_rodada(rodada)
        total_acumulado_excluidos += concluidos

        # Salva o relatório markdown diretamente no Google Drive
        timestamp_slug = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"Relatorio_Rodada_{rodada}_{timestamp_slug}.md"
        caminho_drive = os.path.join(DRIVE_DIR, nome_arquivo)
        
        try:
            with open(caminho_drive, "w", encoding="utf-8") as f:
                f.write(relatorio)
            log_local(f"Relatório salvo com sucesso em: {caminho_drive}")
        except Exception as e:
            log_local(f"ERRO ao salvar relatório no Google Drive: {e}")

        # Decisão de Loop: Paramos se não houver exclusões nem erros/timeouts (tudo limpo!)
        if concluidos == 0 and erros == 0:
            log_local("SUCESSO: Rodada 100% limpa atingida (0 anúncios pendentes de exclusão).")
            
            # Gera relatório consolidado final
            nome_final = f"Relatorio_Final_CONCLUIDO_{timestamp_slug}.md"
            caminho_final = os.path.join(DRIVE_DIR, nome_final)
            conteudo_final = f"""# Relatório Consolidado de Conclusão

A automação de exclusão de anúncios no Mercado Livre foi concluída com sucesso absoluto! Todos os anúncios pendentes listados foram completamente removidos da base.

- **Total de Rodadas Realizadas:** {rodada}
- **Total Acumulado de Anúncios Excluídos:** {total_acumulado_excluidos}
- **Data/Hora de Encerramento:** {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

Todos os anúncios da planilha excluir_2.xlsx agora constam como excluídos com sucesso.
"""
            try:
                with open(caminho_final, "w", encoding="utf-8") as f:
                    f.write(conteudo_final)
                log_local(f"Relatório consolidado de encerramento salvo em: {caminho_final}")
            except Exception as e:
                log_local(f"Erro ao salvar relatório final consolidado: {e}")
            break
        
        # Se restaram anúncios ou ocorreram falhas/timeouts, faz nova rodada
        log_local(f"Ainda há pendências (Excluídos nesta rodada: {concluidos}, Falhas/Timeouts a re-tentar: {erros}).")
        
        if rodada == MAX_RODADAS:
            log_local("AVISO: Limite máximo de rodadas atingido. Encerrando orquestração.")
            break
            
        log_local("Aguardando 15 segundos para estabilização da página do Mercado Livre e do Chrome antes da próxima rodada...")
        time.sleep(15)
        rodada += 1

    log_local("\n=== ORQUESTRADOR FINALIZADO ===")

if __name__ == "__main__":
    main()
