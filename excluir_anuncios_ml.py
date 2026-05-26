import os
import sys
import time
import openpyxl
import csv
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

# Garante suporte completo a caracteres UTF-8 no console do Windows
try:
    sys.stdout.reconfigure(encoding='utf-8')
except AttributeError:
    pass

# Caminhos direcionados para a planilha de não excluídos do Desktop e novo log de finalização
EXCEL_PATH = r"C:\Users\joaos\OneDrive\Desktop\anuncios_nao_excluidos.xlsx"
LISTA_URL = "https://www.mercadolivre.com.br/anuncios/lista"

LOG_DRIVE_PATH = r"G:\Meu Drive\LoselNotebook\LogsSkuScript\status_exclusoes_final.csv"
TRACKER_PATH = r"G:\Meu Drive\LoselNotebook\LogsSkuScript\Acompanhamento_Tarefa_Atual.md"

# == Le todos os IDs de anuncios do relatorio markdown de auditoria (fallback) =====
def ler_anuncios_do_relatorio() -> list[dict]:
    report_path = r"G:\Meu Drive\LoselNotebook\LogsSkuScript\Relatorio_Verificacao_Rapida.md"
    if not os.path.exists(report_path):
        return []
    items = []
    seen = set()
    try:
        with open(report_path, "r", encoding="utf-8") as f:
            for line in f:
                if "ATIVO" in line and "|" in line:
                    parts = [p.strip() for p in line.split("|")]
                    if len(parts) >= 4:
                        id_anuncio = parts[1].replace("`", "").strip()
                        sku = parts[2].replace("`", "").strip()
                        if id_anuncio.startswith("MLB") and id_anuncio not in seen:
                            items.append({"id": id_anuncio, "sku": sku})
                            seen.add(id_anuncio)
    except Exception as e:
        print(f"[WARN] Nao foi possivel ler do relatorio markdown: {e}")
    return items

# == Le todos os IDs de anuncios da planilha ===================================
def ler_anuncios(path: str) -> list[dict]:
    try:
        print(f"[INFO] Lendo planilha: {path}")
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        
        # Obtém cabeçalhos
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        
        col_id = -1
        col_sku = -1
        for idx, header in enumerate(headers):
            if header.lower() == "id anuncio":
                col_id = idx
            elif header.lower() == "sku":
                col_sku = idx
                
        if col_id == -1:
            raise ValueError(f"Coluna 'ID Anuncio' nao encontrada. Cabecalhos: {headers}")
            
        items = []
        seen = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > col_id:
                val_id = str(row[col_id]).strip() if row[col_id] else ""
                val_sku = str(row[col_sku]).strip() if col_sku != -1 and len(row) > col_sku and row[col_sku] else ""
                if val_id and val_id not in seen:
                    items.append({"id": val_id, "sku": val_sku})
                    seen.add(val_id)
                    
        print(f"[PLANILHA] {len(items)} anuncios únicos encontrados.")
        return items
    except PermissionError:
        print(f"[WARN] Planilha {path} esta aberta no Excel e bloqueada. Tentando ler dados diretamente do Relatorio de Verificacao Rapida (Markdown)...")
        items = ler_anuncios_do_relatorio()
        if items:
            print(f"[INFO] {len(items)} anuncios ATIVOS carregados com sucesso do relatorio Markdown!")
            return items
        else:
            raise PermissionError(f"Planilha bloqueada pelo Excel e relatorio Markdown nao pôde ser carregado.")

# == Controle de Log e Re-tentativa ===========================================
def carregar_excluidos_concluidos() -> set[str]:
    excluidos = set()
    if not os.path.exists(LOG_DRIVE_PATH):
        try:
            os.makedirs(os.path.dirname(LOG_DRIVE_PATH), exist_ok=True)
            with open(LOG_DRIVE_PATH, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["data_hora", "id_anuncio", "sku", "status", "detalhe"])
        except Exception as e:
            print(f"[WARN] Nao foi possivel criar o arquivo de log no Google Drive: {e}")
        return excluidos

    try:
        with open(LOG_DRIVE_PATH, "r", newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader, None)  # pula cabeçalho
            for row in reader:
                if len(row) >= 4:
                    id_anuncio = row[1].strip()
                    status = row[3].strip()
                    if status == "CONCLUIDO":
                        excluidos.add(id_anuncio)
    except Exception as e:
        print(f"[ERRO] Falha ao ler log de exclusoes: {e}")
    return excluidos

def registrar_status_item(id_anuncio: str, sku: str, status: str, detalhe: str):
    try:
        os.makedirs(os.path.dirname(LOG_DRIVE_PATH), exist_ok=True)
        with open(LOG_DRIVE_PATH, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                id_anuncio,
                sku,
                status,
                detalhe
            ])
    except Exception as e:
        print(f"[ERRO] Falha ao registrar status do item {id_anuncio} no log: {e}")

# == Atualiza o Dashboard no Obsidian em Tempo Real ============================
def atualizar_dashboard_tempo_real(todos_anuncios: list[dict], item_corrente: dict = None, status_corrente: str = None, detalhe_corrente: str = None):
    excluidos_concluidos = set()
    historico_recentes = []
    
    if os.path.exists(LOG_DRIVE_PATH):
        try:
            with open(LOG_DRIVE_PATH, "r", newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader, None)  # pula cabeçalho
                for row in reader:
                    if len(row) >= 5:
                        data_hora_log = row[0].strip()
                        id_anuncio = row[1].strip()
                        sku = row[2].strip()
                        status = row[3].strip()
                        detalhe = row[4].strip()
                        
                        if status == "CONCLUIDO":
                            excluidos_concluidos.add(id_anuncio)
                        
                        historico_recentes.append({
                            "data_hora": data_hora_log,
                            "id": id_anuncio,
                            "sku": sku,
                            "status": status,
                            "detalhe": detalhe
                        })
        except Exception as e:
            print(f"[WARN] Nao foi possivel ler log para o dashboard: {e}")

    total_anuncios = len(todos_anuncios)
    total_excluidos = len(excluidos_concluidos)
    porcentagem = int((total_excluidos / total_anuncios) * 100) if total_anuncios > 0 else 0
    
    data_hora_agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if total_excluidos >= total_anuncios:
        status_geral_html = '<span style="background-color: #2e7d32; color: #ffffff; padding: 3px 8px; border-radius: 12px; font-weight: 600; font-size: 0.85em; box-shadow: 1px 1px 3px rgba(0,0,0,0.15);">🏆 EXCLUSÃO DE PENDÊNCIAS CONCLUÍDA</span>'
    else:
        status_geral_html = f'<span style="background-color: #1565c0; color: #ffffff; padding: 3px 8px; border-radius: 12px; font-weight: 600; font-size: 0.85em; box-shadow: 1px 1px 3px rgba(0,0,0,0.15);">⚡ EM EXECUÇÃO (LIMPEZA FINAL)</span>'
        
    if item_corrente:
        id_curr = item_corrente["id"]
        sku_curr = item_corrente["sku"]
        if status_corrente == "PROCESSANDO":
            badge_curr = '<span style="background-color: #1565c0; color: white; padding: 1.5px 6px; border-radius: 4px; font-weight: bold; font-size: 0.8em;">⚙️ PROCESSANDO PENDÊNCIA</span>'
            detalhe_curr = "Buscando e excluindo anúncio pendente..."
        elif status_corrente == "CONCLUIDO":
            badge_curr = '<span style="background-color: #2e7d32; color: white; padding: 1.5px 6px; border-radius: 4px; font-weight: bold; font-size: 0.8em;">✅ CONCLUÍDO</span>'
            detalhe_curr = detalhe_corrente
        else:
            badge_curr = '<span style="background-color: #c62828; color: white; padding: 1.5px 6px; border-radius: 4px; font-weight: bold; font-size: 0.8em;">❌ FALHA / LAG</span>'
            detalhe_curr = detalhe_corrente
            
        card_item_atual = f"""
> [!info] **Anúncio em Processamento Atual (Limpeza dos 39 Pendentes)**
> * 🆔 **ID Anúncio:** `{id_curr}`
> * 🏷️ **SKU:** `{sku_curr}`
> * 📊 **Status:** {badge_curr}
> * 💬 **Detalhes:** *{detalhe_curr}*
"""
    else:
        card_item_atual = """
> [!info] **Anúncio em Processamento Atual (Limpeza)**
> * *Nenhum anúncio pendente sendo processado no momento.*
"""

    tabela_linhas = []
    for log in reversed(historico_recentes[-15:]):
        if log["status"] == "CONCLUIDO":
            badge = '<span style="background-color: #2e7d32; color: white; padding: 1px 5px; border-radius: 4px; font-weight: bold; font-size: 0.8em;">CONCLUÍDO</span>'
        else:
            badge = '<span style="background-color: #c62828; color: white; padding: 1px 5px; border-radius: 4px; font-weight: bold; font-size: 0.8em;">FALHA / LAG</span>'
        
        hora_str = log["data_hora"].split(" ")[1] if " " in log["data_hora"] else log["data_hora"]
        tabela_linhas.append(f"| **{hora_str}** | `{log['id']}` | `{log['sku']}` | {badge} | {log['detalhe']} |")
        
    if not tabela_linhas:
        tabela_linhas_str = "| - | - | - | - | Nenhum anúncio processado ainda |"
    else:
        tabela_linhas_str = "\n".join(tabela_linhas)

    md_content = f"""# 👑 Painel de Controle - Exclusão de Anúncios no Mercado Livre

> [!abstract] **Informações do Sistema**
> * 📁 **Planilha de Origem:** `anuncios_nao_excluidos.xlsx` (Desktop)
> * 🎯 **Estratégia:** Busca Cirúrgica por **ID do Anúncio (MLB ID)**
> * 🔄 **Status Geral:** {status_geral_html}
> * 📆 **Última Atualização:** `{data_hora_agora}`

---

## 📊 Progresso da Limpeza dos 39 Anúncios Pendentes

<div style="display: flex; align-items: center; gap: 15px; margin: 15px 0;">
  <div style="flex-grow: 1; background-color: #333333; border-radius: 8px; padding: 4px; border: 1px solid #555555; height: 32px; width: 100%;">
    <div style="width: {porcentagem}%; background: linear-gradient(90deg, #1565c0, #2e7d32); height: 100%; border-radius: 6px; display: flex; align-items: center; justify-content: center; color: white; font-weight: bold; font-family: sans-serif; font-size: 0.9em; box-shadow: inset 0px 1px 2px rgba(255,255,255,0.2);">
      {porcentagem}% ({total_excluidos} / {total_anuncios})
    </div>
  </div>
</div>

---

{card_item_atual}

---

## 🕒 Histórico de Exclusões Recentes (Últimos 15)

<details open>
<summary>Clique para expandir/recolher a lista de histórico</summary>

| Horário | ID Anúncio | SKU | Status | Detalhes / Ação |
| :--- | :--- | :--- | :--- | :--- |
{tabela_linhas_str}

</details>

---

## 🛠️ Próximos Passos Planejados
1. Finalizar a exclusão absoluta de todos os 39 anúncios que restaram ativos na base.
2. Executar uma verificação final rápida para confirmar o status 100% Limpo!
"""

    try:
        os.makedirs(os.path.dirname(TRACKER_PATH), exist_ok=True)
        with open(TRACKER_PATH, "w", encoding="utf-8") as f:
            f.write(md_content)
    except Exception as e:
        print(f"[WARN] Nao foi possivel escrever Acompanhamento_Tarefa_Atual.md: {e}")

# == Exclui o anuncio pelo ID individual =======================================
def excluir_por_id(page, item: dict) -> tuple[str, str]:
    id_anuncio = item["id"]
    sku = item["sku"]
    print(f"\n[ID {id_anuncio}] Buscando anuncio (SKU: {sku})...")

    import urllib.parse
    url_busca = f"https://www.mercadolivre.com.br/anuncios?page=1&sort=DEFAULT&search={urllib.parse.quote(id_anuncio)}"

    # Retry de navegação em sessões longas
    for tentativa_nav in range(3):
        try:
            page.goto(url_busca, wait_until="domcontentloaded", timeout=20000)
            break
        except Exception as e:
            if tentativa_nav < 2:
                print(f"[ID {id_anuncio}] Erro de navegacao (tentativa {tentativa_nav+1}/3): {e} - retentando...")
                time.sleep(3)
            else:
                print(f"[ID {id_anuncio}] Falha ao navegar apos 3 tentativas. Pulando.")
                return "ERRO", "Falha de navegacao"

    time.sleep(2.5)  # Otimizado para 2.5 segundos (velocidade e segurança equilibradas)

    # Localiza botões de ações
    botoes = page.get_by_role("button", name="Ações secundárias").all()
    if not botoes:
        botoes = page.get_by_role("button", name="Acciones secundarias").all()
    if not botoes:
        botoes = page.locator("button.sll-list-row-secondary-actions-trigger").all()

    if not botoes:
        print(f"[ID {id_anuncio}] Nenhum anuncio encontrado na busca. Assumindo como ja excluido.")
        return "CONCLUIDO", "Ja excluido (nao encontrado)"

    print(f"[ID {id_anuncio}] Anuncio encontrado. Clicando no menu de acoes...")

    try:
        botoes[0].scroll_into_view_if_needed()
        botoes[0].click()
        time.sleep(1.5)  # Otimizado para 1.5s (aguarda o popper abrir)

        # Confirma que o popper abriu
        popper = page.locator('[data-testid="popper"]')
        popper.wait_for(state="visible", timeout=6000)

        # Clicar via JS no botão Excluir do popper
        clicou = page.evaluate("""() => {
            const popper = document.querySelector('[data-testid="popper"]');
            if (!popper) return false;
            const btns = Array.from(popper.querySelectorAll('button.andes-list__item-action, button, a, [role="button"], li'));
            const excluir = btns.find(b => b.textContent.trim().includes('Excluir') || b.textContent.trim().includes('Eliminar'));
            if (excluir) { excluir.click(); return true; }
            return false;
        }""")

        if not clicou:
            print(f"[ID {id_anuncio}] Botao Excluir nao encontrado no popper.")
            page.keyboard.press("Escape")
            return "ERRO", "Botao Excluir nao encontrado"

        print(f"[ID {id_anuncio}] Clicou em Excluir. Aguardando modal de confirmacao...")
        time.sleep(1.5)  # Otimizado para 1.5s (aguarda o modal)

        # Confirma a exclusão no modal via JS
        confirmou = page.evaluate("""() => {
            const candidatos = document.querySelectorAll(
                '.andes-modal button.andes-button--loud, ' +
                '.andes-modal button.andes-button--primary, ' +
                '[data-testid="modal"] button.andes-button--loud, ' +
                '[data-testid="modal"] button.andes-button--primary'
            );
            for (const btn of candidatos) {
                if (btn.offsetParent !== null) { // visível
                    btn.click();
                    return 'loud/primary: ' + btn.textContent.trim();
                }
            }
            const popper = document.querySelector('[data-testid="popper"]');
            const todos = document.querySelectorAll('button');
            for (const btn of todos) {
                if (popper && popper.contains(btn)) continue;
                if (!btn.offsetParent) continue;
                const txt = btn.textContent.trim();
                if (txt === 'Excluir' || txt === 'Confirmar' || txt === 'Eliminar' || txt === 'Sim') {
                    btn.click();
                    return 'fallback: ' + txt;
                }
            }
            return null;
        }""")

        if confirmou:
            print(f"[ID {id_anuncio}] Confirmacao clicada ({confirmou}).")
            
            # 1. Verifica se a mensagem verde de confirmação (andes-toast) apareceu
            toast_detectado = False
            toast_text = ""
            print(f"[ID {id_anuncio}] Aguardando mensagem de confirmacao...")
            try:
                toast_loc = page.locator(".andes-toast, .andes-snackbar, [role='status'], .andes-message, [data-testid='toast']").first
                toast_loc.wait_for(state="visible", timeout=6000)
                toast_text = toast_loc.inner_text().strip()
                print(f"[ID {id_anuncio}] Toast detectado: '{toast_text}'")
                
                sucesso_words = ["excluido", "excluído", "excluída", "excluida", "eliminado", "eliminada", "apagado", "apagada", "sucesso", "concluido", "concluído"]
                if any(w in toast_text.lower() for w in sucesso_words):
                    toast_detectado = True
            except Exception as e:
                print(f"[ID {id_anuncio}] Toast de confirmacao nao apareceu dentro de 6s.")
            
            # 2. Recarrega a página de busca para obter a verdade absoluta e dupla confirmação
            print(f"[ID {id_anuncio}] Recarregando pagina para dupla confirmacao...")
            try:
                page.reload(wait_until="domcontentloaded", timeout=12000)
                time.sleep(1.0)
                
                botoes_restantes = page.get_by_role("button", name="Ações secundárias").all()
                if not botoes_restantes:
                    botoes_restantes = page.get_by_role("button", name="Acciones secundarias").all()
                if not botoes_restantes:
                    botoes_restantes = page.locator("button.sll-list-row-secondary-actions-trigger").all()
                
                if not botoes_restantes:
                    print(f"[ID {id_anuncio}] Confirmado: anuncio sumiu da busca! ✨")
                    if toast_detectado:
                        return "CONCLUIDO", f"Excluido e confirmado por toast ('{toast_text}')"
                    else:
                        return "CONCLUIDO", "Excluido (confirmado por desaparecimento na busca)"
                else:
                    print(f"[ID {id_anuncio}] AVISO: anuncio ainda continua visivel na busca apos recarga.")
                    return "ERRO", "Exclusao clicada mas anuncio continua ativo"
            except Exception as e:
                print(f"[ID {id_anuncio}] Erro na dupla confirmacao por recarga: {e}")
                if toast_detectado:
                    # Se o toast confirmou sucesso, assumimos sucesso mesmo com erro de recarga
                    return "CONCLUIDO", f"Excluido (confirmado apenas por toast: '{toast_text}')"
                return "ERRO", f"Erro na confirmacao de recarga: {e}"
        else:
            print(f"[ID {id_anuncio}] Modal de confirmacao nao encontrado.")
            page.keyboard.press("Escape")
            return "ERRO", "Modal de confirmacao nao encontrado"

    except PlaywrightTimeout:
        print(f"[ID {id_anuncio}] TIMEOUT - lag ou delay na resposta da pagina.")
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass
        return "ERRO", "Timeout / Lag na interface"
    except Exception as e:
        print(f"[ID {id_anuncio}] ERRO: {e}")
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass
        return "ERRO", f"Erro inesperado: {e}"

# == Main ======================================================================
def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"[ERRO CRITICO] Planilha {EXCEL_PATH} nao encontrada!")
        sys.exit(1)

    excluidos_ja_concluidos = carregar_excluidos_concluidos()
    todos_anuncios = ler_anuncios(EXCEL_PATH)
    
    anuncios_para_processar = [a for a in todos_anuncios if a["id"] not in excluidos_ja_concluidos]
    
    print(f"\n[RESUMO INICIAL] Total na planilha: {len(todos_anuncios)}")
    print(f"[RESUMO INICIAL] Já concluídos anteriormente: {len(excluidos_ja_concluidos)}")
    print(f"[RESUMO INICIAL] A processar nesta rodada: {len(anuncios_para_processar)}")
    
    # Atualiza o dashboard inicial com o progresso atual
    atualizar_dashboard_tempo_real(todos_anuncios)
    
    if not anuncios_para_processar:
        print("\n[INFO] Todos os anuncios já foram concluídos com sucesso! Nada para fazer.")
        sys.exit(0)

    with sync_playwright() as pw:
        print("\n[CDP] Conectando ao seu Chrome pessoal na porta 9222...")
        try:
            browser = pw.chromium.connect_over_cdp("http://localhost:9222")
            print("[CDP] Conectado com sucesso!")
        except Exception as e:
            print(f"[ERRO CRITICO] Nao foi possivel conectar ao Chrome na porta 9222: {e}")
            print("Certifique-se de que o Chrome foi aberto com depuracao activa usando abrir_chrome_debug.bat")
            sys.exit(1)

        context = browser.contexts[0]
        
        page = None
        for p in context.pages:
            if "mercadolivre.com.br" in p.url:
                page = p
                print(f"[CDP] Usando aba ativa: {p.url}")
                break
        
        if not page:
            page = context.new_page()
            print("[CDP] Abrindo pagina de anuncios...")
            page.goto(LISTA_URL)
        elif "anuncios" not in page.url:
            print("[CDP] Navegando para a pagina de anuncios...")
            page.goto(LISTA_URL)

        print("[INFO] Aguardando pagina de anuncios carregar...")
        try:
            page.wait_for_load_state("domcontentloaded", timeout=15000)
            print("[INFO] Pagina carregada! Iniciando exclusoes...\n")
        except Exception:
            print("[WARN] Timeout aguardando carregamento da pagina de anuncios. Prosseguindo...")

        total_concluidos_nesta_rodada = 0
        total_erros_nesta_rodada = 0

        for i, item in enumerate(anuncios_para_processar, 1):
            print(f"\n=========================================")
            print(f" Anúncio {i}/{len(anuncios_para_processar)}: {item['id']} (SKU: {item['sku']})")
            print(f"=========================================")
            
            # Atualiza o dashboard: Informa o anúncio que começou a ser PROCESSADO!
            atualizar_dashboard_tempo_real(todos_anuncios, item_corrente=item, status_corrente="PROCESSANDO")
            
            status, detalhe = excluir_por_id(page, item)
            registrar_status_item(item["id"], item["sku"], status, detalhe)
            
            if status == "CONCLUIDO":
                total_concluidos_nesta_rodada += 1
            else:
                total_erros_nesta_rodada += 1
                
            # Atualiza o dashboard: Informa o resultado FINAL do anúncio processado!
            atualizar_dashboard_tempo_real(todos_anuncios, item_corrente=item, status_corrente=status, detalhe_corrente=detalhe)
            
            print(f"[RESUMO ITEM] Status: {status} | Detalhe: {detalhe}")
            time.sleep(1.5)  # Atraso de 1.5s entre cada anúncio para estabilidade
            
        print("\n=========================================")
        print(f" RODADA FINALIZADA")
        print(f" Concluídos nesta rodada: {total_concluidos_nesta_rodada}")
        print(f" Erros/Timeouts nesta rodada: {total_erros_nesta_rodada}")
        print("=========================================")
        
        # Limpa o card de item ativo ao final do processamento
        atualizar_dashboard_tempo_real(todos_anuncios)

if __name__ == "__main__":
    main()
