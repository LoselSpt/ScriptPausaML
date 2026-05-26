from __future__ import annotations

import argparse
import csv
import re
import sys
import time
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

BrowserContext = Any
Locator = Any
Page = Any
PlaywrightError = Exception
PlaywrightTimeoutError = TimeoutError
sync_playwright = None


DEFAULT_ML_URL = "https://www.mercadolivre.com.br/anuncios/lista"


@dataclass(frozen=True)
class SkuRow:
    sku: str
    row_number: int
    curve: str


@dataclass(frozen=True)
class RunConfig:
    spreadsheet: Path
    sheet_name: str | None
    sku_column: str | None
    curve_column: str | None
    curve_value: str
    ml_url: str
    profile_dir: Path
    headless: bool
    confirm_pause: bool
    slow_mo_ms: int
    timeout_ms: int
    log_path: Path
    screenshot_dir: Path | None
    wait_for_login: bool


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def display_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def find_column(headers: list[object], requested: str | None, candidates: Iterable[str]) -> int:
    normalized_headers = [normalize_text(header) for header in headers]

    if requested:
        requested_norm = normalize_text(requested)
        if requested_norm in normalized_headers:
            return normalized_headers.index(requested_norm)
        raise ValueError(f"Coluna informada nao encontrada: {requested}")

    candidate_norms = {normalize_text(candidate) for candidate in candidates}
    for index, normalized_header in enumerate(normalized_headers):
        if normalized_header in candidate_norms:
            return index

    readable_headers = ", ".join(display_text(header) for header in headers if display_text(header))
    raise ValueError(
        "Nao encontrei a coluna esperada. "
        f"Cabecalhos encontrados: {readable_headers or '(sem cabecalhos)'}"
    )


def is_curve_match(value: object, expected: str) -> bool:
    raw = display_text(value)
    normalized = normalize_text(raw)
    expected_norm = normalize_text(expected)
    tokens = re.findall(r"[a-z0-9]+", unicodedata.normalize("NFKD", raw).encode("ascii", "ignore").decode().lower())

    if normalized == expected_norm:
        return True

    if expected_norm == "c":
        return raw.strip().upper() == "C" or normalized == "curvac" or "c" in tokens

    return expected_norm in normalized


def read_curve_c_skus(config: RunConfig) -> list[SkuRow]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "Dependencia ausente: openpyxl. Rode: pip install -r requirements.txt"
        ) from exc

    if not config.spreadsheet.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {config.spreadsheet}")

    workbook = load_workbook(config.spreadsheet, data_only=True)
    sheet = workbook[config.sheet_name] if config.sheet_name else workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("A planilha esta vazia.")

    headers = list(rows[0])
    sku_index = find_column(
        headers,
        config.sku_column,
        candidates=("Sku", "SKU", "Skul", "Skull", "Codigo SKU", "Codigo", "Codigo do produto"),
    )
    curve_index = find_column(
        headers,
        config.curve_column,
        candidates=("Curva", "Curva ABC", "Classificacao", "Classificacao ABC", "ABC"),
    )

    selected: list[SkuRow] = []
    seen: set[str] = set()

    for excel_row_number, row in enumerate(rows[1:], start=2):
        sku = display_text(row[sku_index] if sku_index < len(row) else "")
        curve = display_text(row[curve_index] if curve_index < len(row) else "")
        if not sku or not is_curve_match(curve, config.curve_value):
            continue

        dedupe_key = normalize_text(sku)
        if dedupe_key in seen:
            continue

        selected.append(SkuRow(sku=sku, row_number=excel_row_number, curve=curve))
        seen.add(dedupe_key)

    return selected


def first_visible(locator: Locator, timeout_ms: int = 3000) -> Locator | None:
    try:
        count = locator.count()
    except PlaywrightError:
        return None

    for index in range(count):
        item = locator.nth(index)
        try:
            item.wait_for(state="visible", timeout=timeout_ms)
            if item.is_visible():
                return item
        except PlaywrightError:
            continue
    return None


def find_search_input(page: Page, timeout_ms: int) -> Locator:
    selectors = [
        'input[placeholder*="Buscar" i]',
        'input[placeholder*="Pesquisar" i]',
        'input[placeholder*="SKU" i]',
        'input[aria-label*="Buscar" i]',
        'input[aria-label*="Pesquisar" i]',
        'input[type="search"]',
        'input[type="text"]',
        "textarea",
    ]

    deadline = time.monotonic() + timeout_ms / 1000
    while time.monotonic() < deadline:
        for selector in selectors:
            candidate = first_visible(page.locator(selector), timeout_ms=700)
            if candidate:
                return candidate
        time.sleep(0.4)

    raise PlaywrightTimeoutError("Nao encontrei o campo de busca de anuncios no Mercado Livre.")


def maybe_wait_for_login(page: Page, config: RunConfig) -> None:
    try:
        find_search_input(page, timeout_ms=12000)
        return
    except PlaywrightTimeoutError:
        pass

    if not config.wait_for_login:
        raise RuntimeError(
            "Nao encontrei a tela de anuncios. Rode novamente com --aguardar-login "
            "ou informe a URL correta com --url."
        )

    print(
        "\nFaca login no Mercado Livre e navegue ate Vendas > Anuncios. "
        "Quando o campo de busca estiver visivel, pressione ENTER aqui."
    )
    input()
    find_search_input(page, timeout_ms=config.timeout_ms)


def search_sku(page: Page, sku: str, timeout_ms: int) -> None:
    search = find_search_input(page, timeout_ms=timeout_ms)
    search.click(timeout=timeout_ms)

    modifier = "Meta" if sys.platform == "darwin" else "Control"
    search.press(f"{modifier}+A")
    search.fill(sku)
    search.press("Enter")

    try:
        page.wait_for_load_state("networkidle", timeout=timeout_ms)
    except PlaywrightTimeoutError:
        pass

    page.wait_for_timeout(1200)


def visible_text(page: Page) -> str:
    try:
        return page.locator("body").inner_text(timeout=5000)
    except PlaywrightError:
        return ""


def result_count_hint(page: Page) -> str:
    text = visible_text(page)
    match = re.search(r"(\d+)\s+an[u\u00fa]ncios?", text, flags=re.IGNORECASE)
    return match.group(0) if match else ""


def no_results_visible(page: Page) -> bool:
    text = normalize_text(visible_text(page))
    no_result_markers = (
        "naoencontramos",
        "nenhumresultado",
        "naoharesultados",
        "semresultados",
        "0anuncios",
    )
    return any(marker in text for marker in no_result_markers)


def click_select_all(page: Page, timeout_ms: int) -> bool:
    checkbox_candidates = [
        page.get_by_role("checkbox", name=re.compile("Selecionar", re.IGNORECASE)),
        page.locator('label:has-text("Selecionar anuncios") input[type="checkbox"]'),
        page.locator('input[type="checkbox"]').first,
    ]

    for candidate in checkbox_candidates:
        locator = candidate() if callable(candidate) else candidate
        try:
            locator.wait_for(state="visible", timeout=timeout_ms)
            if not locator.is_checked():
                locator.click(timeout=timeout_ms)
            return True
        except PlaywrightError:
            continue

    return False


def click_pause_button(page: Page, timeout_ms: int) -> bool:
    pause_candidates = [
        page.get_by_role("button", name=re.compile(r"^Pausar$", re.IGNORECASE)),
        page.locator('button:has-text("Pausar")'),
        page.locator('[role="button"]:has-text("Pausar")'),
        page.get_by_text(re.compile(r"^Pausar$", re.IGNORECASE)),
    ]

    for candidate in pause_candidates:
        button = first_visible(candidate, timeout_ms=1200)
        if not button:
            continue
        try:
            if button.is_disabled():
                continue
        except PlaywrightError:
            pass

        try:
            button.click(timeout=timeout_ms)
            return True
        except PlaywrightError:
            continue

    return False


def confirm_pause_if_needed(page: Page, timeout_ms: int) -> None:
    page.wait_for_timeout(800)
    labels = (
        r"^Pausar$",
        r"Pausar an[u\u00fa]ncios?",
        r"Confirmar",
        r"Sim,?\s*pausar",
        r"Continuar",
        r"Entendi",
    )

    dialog = first_visible(page.get_by_role("dialog"), timeout_ms=1000)
    search_root = dialog if dialog else page

    for label in labels:
        button = first_visible(
            search_root.get_by_role("button", name=re.compile(label, re.IGNORECASE)),
            timeout_ms=1200,
        )
        if not button:
            continue
        try:
            if button.is_disabled():
                continue
        except PlaywrightError:
            pass

        try:
            button.click(timeout=timeout_ms)
            page.wait_for_timeout(1200)
            return
        except PlaywrightError:
            continue


def take_error_screenshot(page: Page, config: RunConfig, sku: str) -> str:
    if not config.screenshot_dir:
        return ""

    config.screenshot_dir.mkdir(parents=True, exist_ok=True)
    safe_sku = re.sub(r"[^a-zA-Z0-9_-]+", "_", sku).strip("_") or "sku"
    path = config.screenshot_dir / f"erro_{safe_sku}_{datetime.now():%Y%m%d_%H%M%S}.png"
    page.screenshot(path=str(path), full_page=True)
    return str(path)


def pause_sku(page: Page, sku_row: SkuRow, config: RunConfig) -> tuple[str, str]:
    search_sku(page, sku_row.sku, config.timeout_ms)

    if no_results_visible(page):
        return "sem_anuncios", "Nenhum anuncio encontrado para o SKU pesquisado."

    hint = result_count_hint(page)

    if not click_select_all(page, config.timeout_ms):
        return "erro", "Nao consegui localizar o checkbox de selecionar anuncios."

    if not config.confirm_pause:
        return "simulado", f"SKU selecionado para teste. {hint}".strip()

    if not click_pause_button(page, config.timeout_ms):
        return "erro", "Nao consegui clicar no botao Pausar. Verifique se os anuncios ja estao pausados."

    confirm_pause_if_needed(page, config.timeout_ms)

    try:
        page.wait_for_load_state("networkidle", timeout=config.timeout_ms)
    except PlaywrightTimeoutError:
        pass

    return "pausado", f"Comando de pausa enviado. {hint}".strip()


def write_log_header_if_needed(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return

    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(["data_hora", "sku", "linha_excel", "curva", "status", "detalhe", "screenshot"])


def append_log(path: Path, sku_row: SkuRow, status: str, detail: str, screenshot: str = "") -> None:
    with path.open("a", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(
            [
                datetime.now().isoformat(timespec="seconds"),
                sku_row.sku,
                sku_row.row_number,
                sku_row.curve,
                status,
                detail,
                screenshot,
            ]
        )


def run_automation(config: RunConfig) -> int:
    global PlaywrightError, PlaywrightTimeoutError, sync_playwright

    try:
        from playwright.sync_api import (
            Error as ImportedPlaywrightError,
            TimeoutError as ImportedPlaywrightTimeoutError,
            sync_playwright as imported_sync_playwright,
        )
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "Dependencia ausente: playwright. Rode: pip install -r requirements.txt "
            "e depois python -m playwright install chromium"
        ) from exc

    PlaywrightError = ImportedPlaywrightError
    PlaywrightTimeoutError = ImportedPlaywrightTimeoutError
    sync_playwright = imported_sync_playwright

    sku_rows = read_curve_c_skus(config)
    write_log_header_if_needed(config.log_path)

    print(f"SKUs Curva {config.curve_value.upper()} encontrados: {len(sku_rows)}")
    if not sku_rows:
        print("Nada para pausar.")
        return 0

    if not config.confirm_pause:
        print("Modo SIMULACAO ativo. Use --confirmar-pausa para pausar de verdade.")

    with sync_playwright() as playwright:
        context: BrowserContext = playwright.chromium.launch_persistent_context(
            user_data_dir=str(config.profile_dir),
            headless=config.headless,
            slow_mo=config.slow_mo_ms,
            viewport={"width": 1440, "height": 900},
            args=["--start-maximized"],
        )
        page = context.pages[0] if context.pages else context.new_page()
        page.set_default_timeout(config.timeout_ms)

        try:
            page.goto(config.ml_url, wait_until="domcontentloaded", timeout=config.timeout_ms)
            maybe_wait_for_login(page, config)

            for index, sku_row in enumerate(sku_rows, start=1):
                print(f"[{index}/{len(sku_rows)}] Processando SKU {sku_row.sku}...")
                screenshot = ""
                try:
                    status, detail = pause_sku(page, sku_row, config)
                except Exception as exc:  # noqa: BLE001 - this script must keep processing the next SKU.
                    status = "erro"
                    detail = str(exc)
                    try:
                        screenshot = take_error_screenshot(page, config, sku_row.sku)
                    except PlaywrightError:
                        screenshot = ""

                append_log(config.log_path, sku_row, status, detail, screenshot)
                print(f"  -> {status}: {detail}")

        finally:
            context.close()

    print(f"Log salvo em: {config.log_path}")
    return 0


def parse_args() -> RunConfig:
    parser = argparse.ArgumentParser(
        description="Pausa anuncios do Mercado Livre para SKUs classificados como Curva C em uma planilha Excel."
    )
    parser.add_argument("--planilha", required=True, help="Caminho do arquivo .xlsx com as colunas Sku e Curva.")
    parser.add_argument("--aba", help="Nome da aba. Se omitido, usa a primeira aba ativa.")
    parser.add_argument("--coluna-sku", help='Nome exato da coluna de SKU, se nao for "Sku".')
    parser.add_argument("--coluna-curva", help='Nome exato da coluna de curva, se nao for "Curva".')
    parser.add_argument("--valor-curva", default="C", help='Valor da curva que deve ser pausada. Padrao: "C".')
    parser.add_argument("--url", default=DEFAULT_ML_URL, help="URL da pagina de gestao de anuncios do ML.")
    parser.add_argument("--perfil", default="ml-profile", help="Pasta do perfil persistente do navegador.")
    parser.add_argument("--headless", action="store_true", help="Executa sem abrir janela. Use apenas depois de testar.")
    parser.add_argument("--confirmar-pausa", action="store_true", help="Pausa de verdade. Sem isto, roda em simulacao.")
    parser.add_argument("--slow-mo", type=int, default=120, help="Atraso em ms entre acoes do navegador.")
    parser.add_argument("--timeout", type=int, default=20000, help="Timeout em ms para acoes do navegador.")
    parser.add_argument("--log", default="logs/pausa_ml.csv", help="Arquivo CSV de log da execucao.")
    parser.add_argument("--screenshots-erros", default="logs/screenshots", help="Pasta para prints quando houver erro.")
    parser.add_argument(
        "--sem-aguardar-login",
        action="store_true",
        help="Falha imediatamente se o login/tela de anuncios nao estiver pronto.",
    )

    args = parser.parse_args()

    screenshot_dir = Path(args.screenshots_erros) if args.screenshots_erros else None

    return RunConfig(
        spreadsheet=Path(args.planilha).expanduser().resolve(),
        sheet_name=args.aba,
        sku_column=args.coluna_sku,
        curve_column=args.coluna_curva,
        curve_value=args.valor_curva,
        ml_url=args.url,
        profile_dir=Path(args.perfil).expanduser().resolve(),
        headless=args.headless,
        confirm_pause=args.confirmar_pausa,
        slow_mo_ms=args.slow_mo,
        timeout_ms=args.timeout,
        log_path=Path(args.log).expanduser().resolve(),
        screenshot_dir=screenshot_dir.expanduser().resolve() if screenshot_dir else None,
        wait_for_login=not args.sem_aguardar_login,
    )


if __name__ == "__main__":
    try:
        raise SystemExit(run_automation(parse_args()))
    except KeyboardInterrupt:
        print("\nExecucao interrompida pelo usuario.")
        raise SystemExit(130)
